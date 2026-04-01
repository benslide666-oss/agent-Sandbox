import pandas as pd
import datetime
import smtplib
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font

# --- 最新排班映射 (0=周一, ..., 6=周日) ---
WEEKDAY_MAP = {
    0: "陶伟",   # 周一
    1: "张明亮", # 周二
    2: "王风清", # 周三
    3: "王浩",   # 周四
    4: "陶伟",   # 周五
    5: "杨庆华", # 周六
    6: "贡小春"  # 周日
}

def generate_report():
    today = datetime.date.today()
    first_day_this_month = today.replace(day=1)
    last_day_last_month = first_day_this_month - datetime.timedelta(days=1)
    first_day_last_month = last_day_last_month.replace(day=1)
    
    month_num = first_day_last_month.month
    year_num = first_day_last_month.year
    month_str = f"{year_num}年{month_num}月"
    filename = f"{month_str}行政值班统计表.xlsx"

    names_list = ["张明亮", "陶伟", "王风清", "王浩", "贡小春", "杨庆华"]
    counts = {name: 0 for name in names_list}
    
    current_date = first_day_last_month
    while current_date <= last_day_last_month:
        wd = current_date.weekday()
        duty_name = WEEKDAY_MAP[wd]
        if duty_name in counts:
            counts[duty_name] += 1
        current_date += datetime.timedelta(days=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "行政值班统计&加班"
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18

    title_text = f"白兔分院{month_str}行政值班统计"
    ws.merge_cells('A1:B1')
    ws['A1'] = title_text
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws['A2'] = "姓名"
    ws['B2'] = f"{month_num}月值班天数"
    
    row = 3
    for name in names_list:
        ws[f'A{row}'] = name
        ws[f'B{row}'] = counts[name]
        row += 1
    
    ws[f'A{row}'] = "合计"
    ws[f'B{row}'] = sum(counts.values())

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for r in range(1, row + 1):
        for c in range(1, 3):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    row += 2
    ws[f'A{row}'] = "加班人员统计"
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    for name in ["陈雨", "林燕"]:
        ws[f'A{row}'] = name
        ws[f'B{row}'] = ""
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        row += 1
    ws[f'A{row}'] = "合计"
    ws[f'B{row}'] = ""
    ws[f'A{row}'].border = thin_border
    ws[f'B{row}'].border = thin_border

    wb.save(filename)
    return filename, month_str

def send_email(file_path, month_str):
    sender = os.environ.get("SENDER_EMAIL")
    pwd = os.environ.get("SENDER_PWD")
    receiver = os.environ.get("RECEIVER_EMAIL")
    msg = MIMEMultipart()
    msg['Subject'] = f"📅 {month_str}值班报表 (陶/张/王/王/陶/杨/贡)"
    msg['From'], msg['To'] = sender, receiver
    msg.attach(MIMEText(f"您好，附件为{month_str}行政值班统计。", 'plain'))
    with open(file_path, "rb") as f:
        part = MIMEApplication(f.read(), Name=os.path.basename(file_path))
        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        msg.attach(part)
    with smtplib.SMTP_SSL("smtp.qq.com", 465) as s:
        s.login(sender, pwd); s.send_message(msg); s.quit()
    print("✅ 脚本修复完成，报表已发送！")

if __name__ == "__main__":
    f, m = generate_report()
    send_email(f, m)
